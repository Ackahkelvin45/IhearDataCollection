# tasks.py
from celery import shared_task
from .utils import process_audio_file, safe_process_audio_file
from django.core.files import File
import os
from .models import BulkAudioUpload, NoiseDataset
from django.core.exceptions import ValidationError
import logging
from django.db import transaction
from datetime import time
import time as time_module
from .utils import generate_dataset_name, generate_noise_id
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.conf import settings
from celery import group
from django.db.models import Q
from celery.result import AsyncResult

logger = logging.getLogger(__name__)


@shared_task(
    bind=True,
    time_limit=3600 * 6,
    soft_time_limit=3600 * 5,
    ignore_result=False,  # We need results for progress tracking
    result_expires=300,  # Expire results after 5 minutes (300 seconds)
)
def export_with_audio_task(
    self,
    export_history_id,
    folder_structure,
    category_ids,
    export_name,
    filters=None,
    split_count=1,
):
    """Celery task to create export with audio files"""
    from export.models import ExportHistory
    from authentication.models import CustomUser
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import zipfile
    import shutil
    import os
    import math
    from django.db import connection

    # Track temporary files and directories for cleanup
    root_export_dir = None
    zip_path = None
    export_base_dir = None
    use_s3 = False
    temp_export_base = False  # Track if export_base_dir is a temp directory
    s3_upload_successful = False  # Track if S3 upload was successful
    temp_files_created = []  # Track all temporary files created during export
    temp_dirs_created = []  # Track all temporary directories created during export
    all_zip_paths = []  # Track all zip files for split exports

    try:
        # Get export history record
        export_history = ExportHistory.objects.get(id=export_history_id)
        export_history.status = "processing"
        export_history.save()

        user = export_history.user

        # Create export directory - use temp directory if S3 is enabled, otherwise use MEDIA_ROOT
        use_s3 = hasattr(settings, "USE_S3") and settings.USE_S3
        if use_s3:
            # Use temp directory when S3 is enabled since we'll upload and delete
            import tempfile

            temp_dir = tempfile.gettempdir()
            export_base_dir = os.path.join(temp_dir, "exports", f"user_{user.id}")
            temp_export_base = True
        else:
            # Use MEDIA_ROOT for local storage
            media_root = (
                os.path.abspath(settings.MEDIA_ROOT)
                if settings.MEDIA_ROOT
                else os.path.abspath("media")
            )
            export_base_dir = os.path.join(media_root, "exports", f"user_{user.id}")
            temp_export_base = False

        export_base_dir = os.path.abspath(export_base_dir)  # Ensure absolute path
        os.makedirs(export_base_dir, exist_ok=True)
        logger.info(
            f"Export base directory: {export_base_dir} (S3 enabled: {use_s3}, temp: {temp_export_base}, MEDIA_ROOT: {getattr(settings, 'MEDIA_ROOT', 'N/A')})"
        )

        # Build folder structure from JSON
        excel_path = folder_structure.get("excel_path", "")
        audio_path = folder_structure.get("audio_path", "")
        audio_structure_template = export_history.audio_structure_template or ""

        # Create root export directory
        root_export_dir = os.path.join(export_base_dir, f"temp_{export_name}")
        root_export_dir = os.path.abspath(root_export_dir)  # Track absolute path
        os.makedirs(root_export_dir, exist_ok=True)
        temp_dirs_created.append(root_export_dir)  # Track for cleanup

        # Create Excel directory structure (kept on disk)
        excel_dir = os.path.join(root_export_dir, excel_path)
        os.makedirs(excel_dir, exist_ok=True)

        # Validate split_count
        split_count = max(1, min(3, split_count or 1))  # Ensure 1-3 range
        logger.info(f"Export will be split into {split_count} file(s)")

        # Build queryset
        queryset = NoiseDataset.objects.select_related(
            "category",
            "region",
            "community",
            "class_name",
            "subclass",
            "collector",
            "dataset_type",
            "microphone_type",
            "time_of_day",
        )

        # Apply category filter
        if category_ids:
            queryset = queryset.filter(category_id__in=category_ids)

        # Apply other filters (similar to _filtered_noise_queryset)
        if filters:
            if filters.get("search"):
                queryset = queryset.filter(
                    Q(name__icontains=filters["search"])
                    | Q(noise_id__icontains=filters["search"])
                    | Q(description__icontains=filters["search"])
                )
            if filters.get("region"):
                queryset = queryset.filter(region_id=filters["region"])
            if filters.get("community"):
                queryset = queryset.filter(community_id=filters["community"])
            if filters.get("dataset_type"):
                queryset = queryset.filter(dataset_type_id=filters["dataset_type"])
            if filters.get("collector"):
                queryset = queryset.filter(collector_id=filters["collector"])

        total = queryset.count()
        self.update_state(
            state="PROGRESS", meta={"progress": 0, "current": 0, "total": total}
        )

        # Get all dataset IDs and split into parts
        all_dataset_ids = list(queryset.values_list("id", flat=True))

        # Calculate split boundaries
        if split_count == 1:
            dataset_splits = [all_dataset_ids]
        else:
            # Split into equal parts
            split_size = math.ceil(len(all_dataset_ids) / split_count)
            dataset_splits = []
            for i in range(split_count):
                start_idx = i * split_size
                end_idx = min((i + 1) * split_size, len(all_dataset_ids))
                if start_idx < len(all_dataset_ids):
                    dataset_splits.append(all_dataset_ids[start_idx:end_idx])

        logger.info(f"Dataset split sizes: {[len(s) for s in dataset_splits]}")

        # ALL metadata columns
        headers = [
            "Noise ID",
            "Name",
            "Category",
            "Class",
            "Subclass",
            "Region",
            "Community",
            "Collector",
            "Recording Device",
            "Recording Date",
            "Time of Day",
            "Microphone Type",
            "Description",
            "Dataset Type",
            "Audio File Path",
            "Audio File Link",
            "Duration (s)",
            "Sample Rate",
            "Num Samples",
            "RMS Energy",
            "Zero Crossing Rate",
            "Spectral Centroid",
            "Spectral Bandwidth",
            "Spectral Rolloff",
            "Spectral Flatness",
            "Harmonic Ratio",
            "Percussive Ratio",
            "Mean dB",
            "Max dB",
            "Min dB",
            "Std dB",
            "Peak Count",
            "Peak Interval Mean",
            "Dominant Frequency (Hz)",
            "Frequency Range",
            "Event Count",
        ]

        # Style headers
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        # Helper function to build audio subfolder path
        def build_audio_subfolder(dataset, template):
            if not template:
                return ""

            path = template
            path = path.replace(
                "{category}",
                dataset.category.name if dataset.category else "uncategorized",
            )
            path = path.replace(
                "{class}",
                dataset.class_name.name if dataset.class_name else "unclassified",
            )
            path = path.replace(
                "{subclass}",
                dataset.subclass.name if dataset.subclass else "unclassified",
            )
            path = path.replace(
                "{region}", dataset.region.name if dataset.region else "unknown"
            )
            path = path.replace(
                "{community}",
                dataset.community.name if dataset.community else "unknown",
            )
            path = path.replace(
                "{collector}",
                dataset.collector.username if dataset.collector else "unknown",
            )
            path = path.replace(
                "{time_of_day}",
                dataset.time_of_day.name if dataset.time_of_day else "unknown",
            )
            path = path.replace(
                "{recording_date}",
                (
                    dataset.recording_date.strftime("%Y-%m-%d")
                    if dataset.recording_date
                    else "unknown"
                ),
            )
            path = path.replace("{noise_id}", dataset.noise_id or "unknown")
            path = path.replace(
                "{dataset_type}",
                dataset.dataset_type.name if dataset.dataset_type else "unknown",
            )

            import re

            path = re.sub(r'[<>:"|?*]', "_", path)
            path = path.replace("\\", "/").strip("/")

            return path

        # Helper function to create a workbook with headers
        def create_workbook_with_headers():
            wb = Workbook()
            ws = wb.active
            ws.title = "Audio Datasets"
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            return wb, ws

        # Helper function to auto-adjust column widths
        def auto_adjust_columns(ws):
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(header)
                for row in ws.iter_rows(
                    min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num
                ):
                    if row[0].value:
                        max_length = max(max_length, len(str(row[0].value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Process each split part
        all_zip_paths = []
        all_file_sizes = []
        all_download_urls = []
        processed = 0

        for split_idx, split_dataset_ids in enumerate(dataset_splits):
            split_num = split_idx + 1
            split_total = len(dataset_splits)

            # Generate file names with part suffix if split
            if split_count > 1:
                part_export_name = f"{export_name}_part{split_num}of{split_total}"
            else:
                part_export_name = export_name

            logger.info(
                f"Processing split {split_num}/{split_total}: {len(split_dataset_ids)} datasets"
            )

            # Create Excel workbook for this split
            wb, ws = create_workbook_with_headers()

            # Track audio files for this split
            audio_files_to_zip = []

            # Get datasets for this split
            split_queryset = NoiseDataset.objects.filter(
                id__in=split_dataset_ids
            ).select_related(
                "category",
                "region",
                "community",
                "class_name",
                "subclass",
                "collector",
                "dataset_type",
                "microphone_type",
                "time_of_day",
            )

            # Process datasets in this split
            for dataset in split_queryset.iterator(chunk_size=100):
                processed += 1

                # Build audio subfolder (logical path only, not created on disk)
                subfolder = build_audio_subfolder(dataset, audio_structure_template)

                # Download audio file from S3 or local storage (READ-ONLY)
                audio_relative_path = None
                audio_zip_path = None
                if dataset.audio:
                    try:
                        # Get the relative path (name) - this works for both S3 and local storage
                        audio_name = dataset.audio.name

                        if audio_name:
                            original_filename = os.path.basename(audio_name)
                            file_ext = os.path.splitext(original_filename)[1] or ".mp3"
                            audio_filename = f"{dataset.noise_id}{file_ext}"

                            # Logical path of the audio file inside the ZIP archive
                            if subfolder:
                                audio_zip_path = os.path.join(
                                    audio_path, subfolder, audio_filename
                                ).replace("\\", "/")
                            else:
                                audio_zip_path = os.path.join(
                                    audio_path, audio_filename
                                ).replace("\\", "/")

                            # Relative path from Excel file to audio file
                            # Use virtual paths so we don't need the files on disk
                            excel_virtual_dir = os.path.join("root", excel_path)
                            audio_virtual_path = os.path.join("root", audio_zip_path)
                            audio_relative_path = os.path.relpath(
                                audio_virtual_path, excel_virtual_dir
                            ).replace("\\", "/")

                            # Record this audio file so we can stream it directly into the ZIP later
                            audio_files_to_zip.append(
                                {
                                    "audio_field": dataset.audio,
                                    "audio_name": audio_name,
                                    "zip_path": audio_zip_path,
                                    "noise_id": dataset.noise_id,
                                }
                            )
                    except Exception as e:
                        logger.error(
                            f"Error downloading audio for {dataset.noise_id}: {e}",
                            exc_info=True,
                        )

                # Get audio features and analysis (handle missing OneToOne relationships)
                from .models import AudioFeature, NoiseAnalysis

                audio_feature = AudioFeature.objects.filter(
                    noise_dataset=dataset
                ).first()
                noise_analysis = NoiseAnalysis.objects.filter(
                    noise_dataset=dataset
                ).first()

                # Add row to Excel with ALL metadata
                row = [
                    dataset.noise_id or "",
                    dataset.name or "",
                    dataset.category.name if dataset.category else "",
                    dataset.class_name.name if dataset.class_name else "",
                    dataset.subclass.name if dataset.subclass else "",
                    dataset.region.name if dataset.region else "",
                    dataset.community.name if dataset.community else "",
                    dataset.collector.username if dataset.collector else "",
                    dataset.recording_device or "",
                    (
                        dataset.recording_date.strftime("%Y-%m-%d %H:%M:%S")
                        if dataset.recording_date
                        else ""
                    ),
                    dataset.time_of_day.name if dataset.time_of_day else "",
                    dataset.microphone_type.name if dataset.microphone_type else "",
                    dataset.description or "",
                    dataset.dataset_type.name if dataset.dataset_type else "",
                    audio_relative_path or "",
                    (
                        f'=HYPERLINK("{audio_relative_path}", "Open Audio")'
                        if audio_relative_path
                        else ""
                    ),
                    audio_feature.duration if audio_feature else "",
                    audio_feature.sample_rate if audio_feature else "",
                    audio_feature.num_samples if audio_feature else "",
                    audio_feature.rms_energy if audio_feature else "",
                    audio_feature.zero_crossing_rate if audio_feature else "",
                    audio_feature.spectral_centroid if audio_feature else "",
                    audio_feature.spectral_bandwidth if audio_feature else "",
                    audio_feature.spectral_rolloff if audio_feature else "",
                    audio_feature.spectral_flatness if audio_feature else "",
                    audio_feature.harmonic_ratio if audio_feature else "",
                    audio_feature.percussive_ratio if audio_feature else "",
                    noise_analysis.mean_db if noise_analysis else "",
                    noise_analysis.max_db if noise_analysis else "",
                    noise_analysis.min_db if noise_analysis else "",
                    noise_analysis.std_db if noise_analysis else "",
                    noise_analysis.peak_count if noise_analysis else "",
                    noise_analysis.peak_interval_mean if noise_analysis else "",
                    noise_analysis.dominant_frequency if noise_analysis else "",
                    noise_analysis.frequency_range if noise_analysis else "",
                    noise_analysis.event_count if noise_analysis else "",
                ]

                ws.append(row)

                # Update progress every 10 records
                if processed % 10 == 0:
                    progress = int((processed / total) * 100)
                    self.update_state(
                        state="PROGRESS",
                        meta={
                            "progress": progress,
                            "current": processed,
                            "total": total,
                        },
                    )

            # Auto-adjust column widths
            auto_adjust_columns(ws)

            # Save Excel file for this split
            excel_file_path = os.path.join(excel_dir, f"{part_export_name}.xlsx")
            wb.save(excel_file_path)

            # Create ZIP file for this split - ensure absolute path
            zip_path = os.path.join(export_base_dir, f"{part_export_name}.zip")
            zip_path = os.path.abspath(zip_path)  # Ensure absolute path
            temp_files_created.append(zip_path)  # Track for cleanup
            all_zip_paths.append(zip_path)
            logger.info(f"Creating ZIP file at: {zip_path}")

            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
                # 1. Add the Excel file at the configured excel_path inside the ZIP
                excel_arcname = os.path.join(
                    excel_path, f"{part_export_name}.xlsx"
                ).replace("\\", "/")
                zipf.write(excel_file_path, excel_arcname)

                # 2. Stream audio files directly into the ZIP without creating temp copies
                for audio_info in audio_files_to_zip:
                    audio_field = audio_info["audio_field"]
                    audio_name = audio_info["audio_name"]
                    audio_zip_path = audio_info["zip_path"]
                    noise_id = audio_info.get("noise_id") or "unknown"

                    try:
                        # Prefer storage.open (works for both S3 and local)
                        with audio_field.storage.open(audio_name, "rb") as source_file:
                            with zipf.open(audio_zip_path, "w") as dest_file:
                                shutil.copyfileobj(source_file, dest_file)
                        logger.debug(
                            f"Streamed audio for {noise_id} into ZIP at {audio_zip_path}"
                        )
                    except Exception as storage_error:
                        logger.error(
                            f"Failed to read audio from storage for {noise_id}: {storage_error}"
                        )

                        # Fallback: if local storage, try using the filesystem path
                        try:
                            if hasattr(audio_field, "path"):
                                source_path = audio_field.path
                                if os.path.exists(source_path):
                                    with open(source_path, "rb") as source_file:
                                        with zipf.open(
                                            audio_zip_path, "w"
                                        ) as dest_file:
                                            shutil.copyfileobj(source_file, dest_file)
                                    logger.debug(
                                        f"Fallback: streamed audio from local path for {noise_id} at {audio_zip_path}"
                                    )
                                else:
                                    logger.warning(
                                        f"Local audio file not found at {source_path} for {noise_id}"
                                    )
                            else:
                                logger.warning(
                                    f"No local path available for audio file {noise_id}"
                                )
                        except Exception as fallback_error:
                            logger.error(
                                f"Fallback streaming also failed for {noise_id}: {fallback_error}",
                                exc_info=True,
                            )

            # Verify ZIP file exists
            if not os.path.exists(zip_path):
                error_msg = f"ZIP file was not created at expected path: {zip_path}"
                logger.error(error_msg)
                export_history.status = "failed"
                export_history.error_message = error_msg
                export_history.completed_at = timezone.now()
                export_history.save()
                raise Exception(error_msg)

            # Calculate file size for this split
            file_size = os.path.getsize(zip_path)
            all_file_sizes.append(file_size)
            logger.info(
                f"Export ZIP file {split_num}/{split_total} created successfully: {zip_path} ({file_size} bytes)"
            )

        # Use the first file for backward compatibility with single file exports
        zip_path = all_zip_paths[0] if all_zip_paths else None
        file_size = all_file_sizes[0] if all_file_sizes else 0

        # Upload to S3 if S3 is enabled, otherwise keep local
        from django.core.files.storage import default_storage
        from django.core.files import File

        s3_path = None
        download_url = f"/export/download/{export_history.id}/"

        # Check if S3 is enabled
        if hasattr(settings, "USE_S3") and settings.USE_S3:
            try:
                # Log S3 configuration
                logger.info(f"S3 Configuration:")
                logger.info(f"  - USE_S3: {settings.USE_S3}")
                logger.info(
                    f"  - AWS_STORAGE_BUCKET_NAME: {getattr(settings, 'AWS_STORAGE_BUCKET_NAME', 'N/A')}"
                )
                logger.info(
                    f"  - AWS_S3_ENDPOINT_URL: {getattr(settings, 'AWS_S3_ENDPOINT_URL', 'N/A')}"
                )
                logger.info(f"  - MEDIA_URL: {getattr(settings, 'MEDIA_URL', 'N/A')}")

                # Upload all ZIP files to S3
                for idx, current_zip_path in enumerate(all_zip_paths):
                    current_file_name = os.path.basename(current_zip_path)
                    s3_relative_path = f"exports/user_{user.id}/{current_file_name}"

                    logger.info(
                        f"Uploading export ZIP {idx + 1}/{len(all_zip_paths)} to S3"
                    )
                    logger.info(f"  - S3 relative path: {s3_relative_path}")
                    logger.info(f"  - Full S3 path will be: media/{s3_relative_path}")
                    logger.info(f"  - ZIP file size: {all_file_sizes[idx]} bytes")
                    logger.info(f"  - ZIP file path: {current_zip_path}")

                    # Verify file exists before upload
                    if not os.path.exists(current_zip_path):
                        raise Exception(
                            f"ZIP file does not exist at {current_zip_path}"
                        )

                    with open(current_zip_path, "rb") as zip_file:
                        # Use Django's default storage to upload
                        # The storage backend will prepend the 'media' location
                        current_s3_path = default_storage.save(
                            s3_relative_path, File(zip_file, name=current_file_name)
                        )
                        logger.info(
                            f"Successfully uploaded to S3. Storage path: {current_s3_path}"
                        )

                        if idx == 0:
                            s3_path = current_s3_path  # Track first file for backward compatibility

                    # Get the S3 URL - the storage backend should provide the full URL
                    try:
                        s3_url = default_storage.url(current_s3_path)
                        all_download_urls.append(s3_url)
                        if idx == 0:
                            download_url = s3_url
                        logger.info(f"S3 URL for export part {idx + 1}: {s3_url}")
                    except Exception as url_error:
                        logger.warning(
                            f"Could not get S3 URL from storage: {str(url_error)}",
                            exc_info=True,
                        )
                        # Fallback: construct URL manually
                        if hasattr(settings, "MEDIA_URL"):
                            # MEDIA_URL already includes the bucket and media path
                            constructed_url = (
                                f"{settings.MEDIA_URL.rstrip('/')}/{s3_relative_path}"
                            )
                            all_download_urls.append(constructed_url)
                            if idx == 0:
                                download_url = constructed_url
                            logger.info(
                                f"Constructed S3 URL manually: {constructed_url}"
                            )
                        else:
                            fallback_url = (
                                f"/export/download/{export_history.id}/?part={idx + 1}"
                            )
                            all_download_urls.append(fallback_url)
                            if idx == 0:
                                download_url = f"/export/download/{export_history.id}/"
                            logger.warning(
                                "Could not construct S3 URL, falling back to local download"
                            )

                    # Verify the file exists in S3
                    try:
                        if default_storage.exists(current_s3_path):
                            logger.info(
                                f"Verified file exists in S3: {current_s3_path}"
                            )
                        else:
                            logger.error(
                                f"File does not exist in S3 after upload: {current_s3_path}"
                            )
                    except Exception as verify_error:
                        logger.warning(
                            f"Could not verify file in S3: {str(verify_error)}"
                        )

                s3_upload_successful = True

            except Exception as e:
                logger.error(f"Failed to upload export to S3: {str(e)}", exc_info=True)
                # Continue with local file if S3 upload fails
                logger.warning("Falling back to local file storage")
                s3_path = None
                s3_upload_successful = False
                # Build local download URLs
                all_download_urls = []
                for idx in range(len(all_zip_paths)):
                    all_download_urls.append(
                        f"/export/download/{export_history.id}/?part={idx + 1}"
                    )
        else:
            # Local storage - build download URLs
            for idx in range(len(all_zip_paths)):
                if split_count > 1:
                    all_download_urls.append(
                        f"/export/download/{export_history.id}/?part={idx + 1}"
                    )
                else:
                    all_download_urls.append(f"/export/download/{export_history.id}/")

        # Update export history
        export_history.status = "completed"
        export_history.download_url = download_url
        export_history.file_size = file_size
        export_history.total_files = total
        export_history.completed_at = timezone.now()

        # Store split file information
        if split_count > 1:
            export_history.download_urls = all_download_urls
            export_history.file_sizes = all_file_sizes

        export_history.save()

        logger.info(
            f"Export {export_history.id} marked as completed. Files: {len(all_zip_paths)}, Total size: {sum(all_file_sizes)} bytes"
        )

        # Return minimal result - all data is already stored in ExportHistory model
        # This prevents Redis from storing large result data
        # The download_url, file_size, etc. are already in the database
        return {
            "status": "completed",
            "export_id": export_history.id,
        }

    except Exception as e:
        logger.error(f"Export task failed: {str(e)}", exc_info=True)

        # Update export history with error
        try:
            export_history = ExportHistory.objects.get(id=export_history_id)
            export_history.status = "failed"
            export_history.error_message = str(e)
            export_history.completed_at = timezone.now()
            export_history.save()
        except:
            pass

        raise

    finally:
        # Clean up all temporary files and directories
        logger.info("Starting cleanup of temporary files and directories...")

        # Track all files/dirs to clean up
        cleanup_items = []

        # 1. Clean up root export directory (temporary folder with Excel and audio files)
        if root_export_dir and os.path.exists(root_export_dir):
            cleanup_items.append(("directory", root_export_dir))

        # 2. Clean up ZIP files
        # Delete ZIP files if:
        # - S3 upload was successful (files are now in S3, local copies not needed)
        # - Using temp directory (should always clean up temp files)
        # - S3 is enabled but upload failed (files should be cleaned up, user can retry)
        should_delete_zip = (
            s3_upload_successful
            or temp_export_base  # S3 upload succeeded, files are in S3
            or (  # Using temp directory
                use_s3 and not s3_upload_successful
            )  # S3 enabled but upload failed
        )

        # Clean up all zip files (for split exports)
        for current_zip_path in all_zip_paths:
            if current_zip_path and os.path.exists(current_zip_path):
                if should_delete_zip:
                    cleanup_items.append(("file", current_zip_path))
                else:
                    # Only keep ZIP file if using local storage (S3 disabled) and export succeeded
                    logger.info(f"Keeping ZIP file for local storage: {current_zip_path}")

        # 3. Clean up export_base_dir if it's a temp directory
        if export_base_dir and temp_export_base and os.path.exists(export_base_dir):
            cleanup_items.append(("directory", export_base_dir))

        # 4. Clean up any other tracked temporary files/directories
        for temp_file in temp_files_created:
            if (
                temp_file
                and os.path.exists(temp_file)
                and temp_file
                not in [item[1] for item in cleanup_items if item[0] == "file"]
            ):
                cleanup_items.append(("file", temp_file))

        for temp_dir in temp_dirs_created:
            if (
                temp_dir
                and os.path.exists(temp_dir)
                and temp_dir
                not in [item[1] for item in cleanup_items if item[0] == "directory"]
            ):
                cleanup_items.append(("directory", temp_dir))

        # Perform cleanup
        for item_type, item_path in cleanup_items:
            try:
                if item_type == "directory":
                    if os.path.exists(item_path):
                        shutil.rmtree(item_path)
                        logger.info(f"✓ Cleaned up temporary directory: {item_path}")
                elif item_type == "file":
                    if os.path.exists(item_path):
                        os.remove(item_path)
                        logger.info(f"✓ Cleaned up temporary file: {item_path}")
            except PermissionError as e:
                logger.error(
                    f"✗ Permission denied removing {item_type} {item_path}: {str(e)}"
                )
            except FileNotFoundError:
                logger.debug(f"File/directory already removed: {item_path}")
            except Exception as e:
                logger.error(
                    f"✗ Failed to remove {item_type} {item_path}: {str(e)}",
                    exc_info=True,
                )

        # Additional cleanup: Remove any leftover temp files/directories in export directories
        try:
            if export_base_dir and os.path.exists(export_base_dir):
                # Clean up any remaining temp files/directories older than 1 hour
                current_time = time_module.time()
                one_hour_ago = current_time - 3600

                for root, dirs, files in os.walk(export_base_dir):
                    # Clean up old temp files
                    for file in files:
                        if (
                            file.startswith("temp_")
                            or file.endswith(".tmp")
                            or file.endswith(".zip")
                        ):
                            temp_file_path = os.path.join(root, file)
                            try:
                                # Delete if file is older than 1 hour
                                if os.path.getmtime(temp_file_path) < one_hour_ago:
                                    os.remove(temp_file_path)
                                    logger.info(
                                        f"Removed old temp file: {temp_file_path}"
                                    )
                            except Exception as e:
                                logger.warning(
                                    f"Could not remove temp file {temp_file_path}: {str(e)}"
                                )

                    # Clean up old temp directories
                    for dir_name in dirs:
                        if dir_name.startswith("temp_"):
                            temp_dir_path = os.path.join(root, dir_name)
                            try:
                                # Delete if directory is older than 1 hour
                                if os.path.getmtime(temp_dir_path) < one_hour_ago:
                                    shutil.rmtree(temp_dir_path)
                                    logger.info(
                                        f"Removed old temp directory: {temp_dir_path}"
                                    )
                            except Exception as e:
                                logger.warning(
                                    f"Could not remove temp directory {temp_dir_path}: {str(e)}"
                                )
        except Exception as e:
            logger.warning(f"Error during additional temp file cleanup: {str(e)}")

        logger.info("✓ Cleanup of temporary files and directories completed")


@shared_task
def process_audio_task(noise_dataset_id):
    try:
        instance = NoiseDataset.objects.get(id=noise_dataset_id)
        success = safe_process_audio_file(instance)
        if not success:
            logger.warning(
                f"Audio processing failed for NoiseDataset {noise_dataset_id}"
            )
    except NoiseDataset.DoesNotExist:

        logging.warning(f"NoiseDataset with ID {noise_dataset_id} not found.")


@shared_task
def check_task_revocation(task_id):
    """Check if a task has been revoked - separate from main processing to avoid Numba issues"""
    try:

        result = AsyncResult(task_id)
        return result.revoked()
    except Exception:
        return False


@shared_task(bind=True, time_limit=3600 * 6, soft_time_limit=3600 * 5)
def bulk_reprocess_audio_analysis(self, dataset_ids, user_id=None):
    """
    Bulk reprocess audio analysis for multiple datasets with progress tracking
    Optimized for large datasets (400+ items)
    """
    total_datasets = len(dataset_ids)
    processed_count = 0
    failed_count = 0
    failed_datasets = []

    # Limit failed datasets list to prevent memory issues
    MAX_FAILED_DETAILS = 100

    logger.info(f"Starting bulk reprocessing for {total_datasets} datasets")

    # Update task state
    self.update_state(
        state="PROGRESS",
        meta={
            "current": 0,
            "total": total_datasets,
            "status": "Starting reprocessing...",
            "processed": 0,
            "failed": 0,
            "failed_datasets": [],
        },
    )

    try:
        # Process in smaller batches to prevent memory issues
        batch_size = 50
        for batch_start in range(0, total_datasets, batch_size):
            batch_end = min(batch_start + batch_size, total_datasets)
            batch_ids = dataset_ids[batch_start:batch_end]

            logger.info(
                f"Processing batch {batch_start//batch_size + 1}: datasets {batch_start+1}-{batch_end}"
            )

            for i, dataset_id in enumerate(batch_ids):
                global_index = batch_start + i
                dataset = None  # Initialize dataset variable

                try:
                    # Get the dataset with select_related to reduce DB queries
                    try:
                        dataset = NoiseDataset.objects.select_related(
                            "collector", "category", "region", "community"
                        ).get(id=dataset_id)
                    except NoiseDataset.DoesNotExist:
                        logger.warning(f"Dataset {dataset_id} not found, skipping")
                        failed_count += 1
                        if len(failed_datasets) < MAX_FAILED_DETAILS:
                            failed_datasets.append(
                                {"id": dataset_id, "error": "Dataset not found"}
                            )
                        continue

                    # Check if dataset has audio file
                    if not dataset.audio:
                        logger.warning(
                            f"Dataset {dataset_id} has no audio file, skipping"
                        )
                        failed_count += 1
                        if len(failed_datasets) < MAX_FAILED_DETAILS:
                            failed_datasets.append(
                                {
                                    "id": dataset_id,
                                    "noise_id": dataset.noise_id,
                                    "error": "No audio file",
                                }
                            )
                        continue

                    # Process the audio file
                    logger.info(
                        f"Processing dataset {dataset_id} ({global_index+1}/{total_datasets})"
                    )
                    success = safe_process_audio_file(dataset)

                    if success:
                        processed_count += 1
                        logger.info(f"Successfully processed dataset {dataset_id}")
                    else:
                        failed_count += 1
                        logger.warning(f"Failed to process dataset {dataset_id}")
                        if len(failed_datasets) < MAX_FAILED_DETAILS:
                            failed_datasets.append(
                                {
                                    "id": dataset_id,
                                    "noise_id": (
                                        getattr(dataset, "noise_id", "Unknown")
                                        if dataset
                                        else "Unknown"
                                    ),
                                    "error": "Audio processing failed",
                                }
                            )

                except Exception as e:
                    logger.error(f"Failed to process dataset {dataset_id}: {str(e)}")
                    failed_count += 1
                    if len(failed_datasets) < MAX_FAILED_DETAILS:
                        failed_datasets.append(
                            {
                                "id": dataset_id,
                                "noise_id": (
                                    getattr(dataset, "noise_id", "Unknown")
                                    if dataset
                                    else "Unknown"
                                ),
                                "error": str(e)[:200],  # Limit error message length
                            }
                        )

                # Update progress less frequently for large datasets
                if (global_index + 1) % 10 == 0 or global_index + 1 == total_datasets:
                    progress_percentage = int(
                        ((global_index + 1) / total_datasets) * 100
                    )
                    self.update_state(
                        state="PROGRESS",
                        meta={
                            "current": global_index + 1,
                            "total": total_datasets,
                            "status": f"Processing dataset {global_index+1} of {total_datasets} ({progress_percentage}%)",
                            "processed": processed_count,
                            "failed": failed_count,
                            "failed_datasets": failed_datasets,
                            "progress_percentage": progress_percentage,
                        },
                    )

    except Exception as e:
        logger.error(f"Critical error in bulk reprocessing: {str(e)}")
        # Return partial results
        result = {
            "current": processed_count + failed_count,
            "total": total_datasets,
            "status": f"Failed with error: {str(e)}",
            "processed": processed_count,
            "failed": failed_count,
            "failed_datasets": failed_datasets,
            "progress_percentage": int(
                ((processed_count + failed_count) / total_datasets) * 100
            ),
            "final_status": "failed",
            "error": str(e),
        }
        return result

    # Final state
    final_status = "completed"
    if failed_count > 0 and processed_count == 0:
        final_status = "failed"
    elif failed_count > 0:
        final_status = "completed_with_errors"

    result = {
        "current": total_datasets,
        "total": total_datasets,
        "status": f"Completed: {processed_count} processed, {failed_count} failed",
        "processed": processed_count,
        "failed": failed_count,
        "failed_datasets": failed_datasets,
        "progress_percentage": 100,
        "final_status": final_status,
    }

    logger.info(
        f"Bulk reprocessing completed: {processed_count} processed, {failed_count} failed"
    )
    return result


@shared_task(bind=True)
def process_bulk_upload(self, bulk_upload_id, file_paths, user_id):
    try:
        User = get_user_model()
        user = User.objects.get(id=user_id)

        bulk_upload = BulkAudioUpload.objects.get(id=bulk_upload_id)
        bulk_upload.status = "processing"
        bulk_upload.save()

        # Validate metadata before processing
        try:
            bulk_upload.clean_metadata()
        except ValidationError as e:
            bulk_upload.status = "failed"
            bulk_upload.save()
            logger.error(
                f"Metadata validation failed for bulk upload {bulk_upload_id}: {e}"
            )
            return {
                "bulk_upload_id": bulk_upload_id,
                "processed": 0,
                "failed": len(file_paths),
                "error": str(e),
            }

        metadata = bulk_upload.metadata

        for i, file_path in enumerate(file_paths):
            try:
                # Check for cancellation
                bulk_upload.refresh_from_db()
                if getattr(bulk_upload, "status", "").lower() == "cancelled":
                    logger.info(
                        f"Bulk upload {bulk_upload_id} cancelled. Stopping processing."
                    )
                    break
                # Validate file exists and is accessible
                if not os.path.exists(file_path):
                    raise FileNotFoundError(f"File not found at path: {file_path}")

                if not os.access(file_path, os.R_OK):
                    raise PermissionError(f"No read permission for file: {file_path}")

                # Create NoiseDataset instance
                # Parse recording_date to datetime
                recording_dt = metadata["recording_date"]
                if isinstance(recording_dt, str):
                    recording_dt_parsed = parse_datetime(recording_dt)
                    if recording_dt_parsed and timezone.is_naive(recording_dt_parsed):
                        recording_dt_parsed = timezone.make_aware(recording_dt_parsed)
                else:
                    recording_dt_parsed = recording_dt

                noise_dataset = NoiseDataset(
                    collector=user,
                    description=metadata.get("description"),
                    region_id=metadata["region_id"],
                    category_id=metadata["category_id"],
                    time_of_day_id=metadata["time_of_day_id"],
                    community_id=metadata["community_id"],
                    class_name_id=metadata["class_name_id"],
                    subclass_id=metadata.get("subclass_id"),
                    microphone_type_id=metadata.get("microphone_type_id"),
                    recording_date=recording_dt_parsed,
                    recording_device=metadata["recording_device"],
                )

                # Generate IDs and names
                noise_dataset.noise_id = generate_noise_id(user)
                noise_dataset.name = generate_dataset_name(noise_dataset)

                # Validate model before saving (excluding audio field)
                noise_dataset.full_clean(exclude=["audio"])

                # Save the file with proper naming: rename to noise_id + original ext
                original_name = os.path.basename(file_path)
                with open(file_path, "rb") as f:
                    # Derive extension
                    ext = os.path.splitext(original_name)[1] or ""
                    final_name = f"{noise_dataset.noise_id}{ext}"
                    noise_dataset.audio.save(final_name, File(f))

                noise_dataset.save()

                with transaction.atomic():
                    bulk_upload.processed_files += 1
                    bulk_upload.save()

                # Update task state
                self.update_state(
                    state="PROGRESS",
                    meta={
                        "current": i + 1,
                        "total": len(file_paths),
                        "bulk_upload_id": bulk_upload_id,
                    },
                )

            except Exception as e:
                logger.error(
                    f"Failed to process file {file_path}: {str(e)}", exc_info=True
                )
                with transaction.atomic():
                    bulk_upload.failed_files += 1
                    bulk_upload.save()
                continue
            finally:
                # Clean up temporary file
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                except Exception as e:
                    logger.error(f"Failed to remove temp file {file_path}: {str(e)}")

        # Update final status
        if bulk_upload.status == "cancelled":
            # Clean up any remaining files for the user
            try:
                upload_dir = os.path.join(
                    settings.SHARED_UPLOADS_DIR, f"user_{user.id}"
                )
                for p in os.listdir(upload_dir):
                    full = os.path.join(upload_dir, p)
                    if os.path.isfile(full):
                        os.remove(full)
            except Exception:
                pass
        else:
            if bulk_upload.failed_files == 0:
                bulk_upload.status = "completed"
            elif bulk_upload.processed_files > 0:
                bulk_upload.status = "completed_with_errors"
            else:
                bulk_upload.status = "failed"

        bulk_upload.save()

        return {
            "bulk_upload_id": bulk_upload_id,
            "processed": bulk_upload.processed_files,
            "failed": bulk_upload.failed_files,
        }

    except Exception as e:
        logger.error(
            f"Critical error processing bulk upload {bulk_upload_id}: {str(e)}",
            exc_info=True,
        )
        try:
            bulk_upload.status = "failed"
            bulk_upload.save()
        except Exception:
            pass
        return {
            "bulk_upload_id": bulk_upload_id,
            "processed": bulk_upload.processed_files,
            "failed": bulk_upload.failed_files,
            "error": str(e),
        }


@shared_task
def cleanup_shared_uploads(days_old=1):
    shared_dir = getattr(settings, "SHARED_UPLOADS_DIR", "/shared_uploads")
    now = time.time()
    cutoff = now - (days_old * 86400)

    for filename in os.listdir(shared_dir):
        filepath = os.path.join(shared_dir, filename)
        if os.path.isfile(filepath):
            file_time = os.path.getmtime(filepath)
            if file_time < cutoff:
                try:
                    os.remove(filepath)
                except Exception as e:
                    logger.error(f"Failed to remove {filepath}: {str(e)}")
